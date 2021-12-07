from openpyxl import load_workbook
from tkinter import *
import random


wb = load_workbook('TbP.xlsx')
wr = load_workbook('Pgr.xlsx')
WriteInP = wr['Plasatori']
ReadFromP = wb['DP']
ReadFromBA = wb['DBA']
ReadFromBO = wb['DBO']

Letter = ['C','D','E','F','G','H','I']
Vineri = []
Sambata = []
Duminica = []
Luni = []
Marti = []
Miercuri = []
Joi = []
Disp = []
Days = [Vineri, Sambata, Duminica, Luni, Marti, Miercuri, Joi]


#Citeste din fisierul excel cu preferinte
def Read(ReadFrom):
    for i in range(6,98,6):
        rand = []
        if ReadFrom['D' + str(i-3)].value != 'Nume':
            rand.append(ReadFrom['D'+str(i-3)].value)
            for j in range(2):
                for k in Letter:
                    poz = k + str(i+j)
                    if ReadFrom[poz].value == 'X':
                        rand.append(ReadFrom[poz].value)
                    else:
                        rand.append('O')
            rand.append(0)
            Disp.append(rand)
    return Disp
def Minim(j):
    minim = Disp[7][15]
    for i in range(len(Disp)):
        if Disp[i][j] == 'O':
            if Disp[i][15] < minim:
                minim = Disp[i][15]
    return minim
def WriteInPl(WriteIn):
    l1 = [5,14,26,35]
    l2 = [14,26,35]
    k = 0
    p = 0
    for i in range(7):
        for j in range(len(Days[i])):
            if i % 2 == 0:
                poz = 'H' + str(l1[k]+j)
                WriteIn[poz] = Days[i][j]
            else:
                poz = 'D' + str(l2[p]+j)
                WriteIn[poz] = Days[i][j]
        if i % 2 == 0:
            k += 1
        else:
            p += 1
    wr.save('Pgr.xlsx')
def ComplPDi(Day, j):
    for i in range(len(Disp)):
            if Disp[i][j] == 'O' and Disp[i][0] not in Day and len(Day) < 2:
                if Disp[i][15] <= Minim(j):
                    Disp[i][j] = 'X'
                    Day.append(Disp[i][0])
                    Disp[i][15] += 1
            if len(Day) == 2:
                return Day
def ComplPDu(Day, j):
        for i in range(len(Disp)):
                if Disp[i][j] == 'O' and Disp[i][0] not in Day and len(Day) < 6:
                    if Disp[i][15] <= Minim(j) + 1:
                        Disp[i][j] = 'X'
                        Day.append(Disp[i][0])
                        Disp[i][15] += 1
                else:
                    return Day
                if len(Day) == 6:
                    return Day
def ComplPlasatori():
    Read(ReadFromP)
    NrTure = 0
    random.shuffle(Disp)
    while NrTure < 40:
        NrTure = 0
        j = 1
        for Day in Days:
            ComplPDi(Day,j)
            random.shuffle(Disp)
            j += 1
        j = 8
        for Day in Days:
            ComplPDu(Day,j)
            random.shuffle(Disp)
            j += 1
        for i in range(len(Disp)):
            NrTure += Disp[i][15]
    Done.grid(row = 2, column = 0)
    NrTureAfis = Label(root, text=NrTure)
    NrTureAfis.grid(row = 3, column = 0)
    for i in range(len(Disp)):
        print(Disp[i][0], ':',Disp[i][15])
    WriteInPl(WriteInP)


root = Tk()
Plasatori = Label(root, text='Plasatori')
Done = Label(root, text='Done')
ButtonPlasatori = Button(root, text = 'Generare Program Plasatori', padx = 15 ,pady = 5, command = ComplPlasatori)


Plasatori.grid(row = 0, column = 0)
ButtonPlasatori.grid(row = 1, column = 0)

root.mainloop()
